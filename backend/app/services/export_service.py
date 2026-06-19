"""
app/services/export_service.py
──────────────────────────────────────────────────────────────────────────────
Excel export generation service.

For small exports (<= 5000 rows): generate in-memory and return bytes.
For large exports (> 5000 rows): this is called by the ARQ worker asynchronously.

Uses openpyxl for XLSX generation with proper formatting.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.core.logging import get_logger
from app.models.supplier import Supplier

logger = get_logger(__name__)

# ── Excel styling constants ───────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
ALT_ROW_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")

COLUMNS = [
    ("No", 5),
    ("ID", 38),
    ("Nama Perusahaan", 40),
    ("NPWP", 20),
    ("Nama PIC", 30),
    ("Telepon PIC", 18),
    ("Email PIC", 30),
    ("Alamat", 50),
    ("Kota", 20),
    ("Provinsi", 20),
    ("Kategori", 20),
    ("Status", 18),
    ("Catatan", 40),
    ("Dibuat Oleh (Telegram)", 25),
    ("Tanggal Dibuat", 22),
    ("Tanggal Update", 22),
]


def generate_supplier_excel(suppliers: list[Supplier]) -> bytes:
    """
    Generate a formatted XLSX file from a list of Supplier ORM instances.

    Returns:
        Raw XLSX bytes ready to upload to MinIO or stream to client.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Supplier Data"

    # ── Header row ────────────────────────────────────────────────────────────
    headers = [col[0] for col in COLUMNS]
    col_widths = [col[1] for col in COLUMNS]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"  # Freeze header row

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, supplier in enumerate(suppliers, start=2):
        is_alt_row = row_idx % 2 == 0

        row_data = [
            row_idx - 1,  # No.
            str(supplier.id),
            supplier.company_name,
            supplier.npwp_number or "",
            supplier.pic_name,
            supplier.pic_phone,
            supplier.pic_email or "",
            supplier.address or "",
            supplier.city or "",
            supplier.province or "",
            supplier.category or "",
            supplier.status.value if supplier.status else "",
            supplier.notes or "",
            supplier.submitted_by_telegram_username or "",
            _format_dt(supplier.created_at),
            _format_dt(supplier.updated_at),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if is_alt_row:
                cell.fill = ALT_ROW_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    # ── Metadata ──────────────────────────────────────────────────────────────
    wb.properties.creator = "Supplier CRM"
    wb.properties.description = (
        f"Supplier export generated at "
        f"{datetime.now(tz=timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    )

    # ── Serialize to bytes ────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    content = buffer.read()

    logger.info(
        "excel_generated",
        row_count=len(suppliers),
        size_bytes=len(content),
    )

    return content


def _format_dt(dt: datetime | None) -> str:
    """Format a datetime to ISO string for Excel cells."""
    if dt is None:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M:%S")
